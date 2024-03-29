#!/usr/bin/env python
# encoding: utf-8
# @author: lishaogang
# @file: test.py
# @time: 2020/10/20 0020 18:59
# @desc: 实验测试MODE/D-DE算法及其改进
import math

import openpyxl
import numpy as np
# from utils.benchmarks import *
# from moeadCoDE import MOEADCODE
# from moeaD import MOEAD
# from moeadSaDE import MOEADSADE
# from moeadde import MOEADDE
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

from utils.drawResults import draw_box2, draw_box
from utils.fileProcess import savePareto2Txt, readPareto4Txt, saveArray2Excel, saveRes2Excel, write2file
from utils.igd import get_igd
from utils.pfget import get_pflist
from utils.referencePoint import get_referencepoint, get_referencepoint2
from utils.common import extract_info,draw_scatter3D, draw_igd, hv_count

from config import *

def problems_test(nrun, MODEL, draw, s2f=False):
    '''
    一系列函数问题的测试
    :return:
    '''
    results = {}
    for id in range(7):
        print('DTLZ{} starting……'.format(id+1))
        # problem_test(problem=problems[id], draw=False)
        pf = get_pflist(f'pf_files/n10000/DTLZ{id+1}.txt')
        igds, hvs = n_run(nrun, MODEL, problems[id], pf, draw=draw, s2f=s2f)
        results[f'DTLZ{id+1}'] = [igds, hvs]
    for id in range(9):
        print('WFG{} starting……'.format(id + 1))
        pf = get_pflist(f"pf_files/wfg-pf/WFG{id+1}.3D.pf")
        igds, hvs = n_run(nrun, MODEL, problems[id+7], pf, draw=draw, s2f=s2f)
        results[f'WFG{id + 1}'] = [igds, hvs]
    return results

def problem_test(MODEL, problem, pf, reference_point, draw=True, s2f=False):
    '''
    单个问题测试
    :param problem:
    :param draw:时候画图
    :param s2f:是否保存帕累托前沿到文件中
    :return:
    '''
    model = MODEL(pf=pf, problem=problem)
    model.execute()
    pops, x, y, z = extract_info(model)
    # reference_point = get_referencepoint(pops)
    igd = get_igd(model.pareto_front, pops)  # 计算反世代距离IGD
    hv_score = hv_count(model, reference_point)
    # print('hyper volume is {}'.format(hv_score))
    # print('inverted generational distance is {}'.format(igd))
    if draw:
        draw_scatter3D(model.pname, hv_score, igd, reference_point, x, y, z, model.pareto_front)
        # draw_igd(distances, model)
    if s2f:
        savePareto2Txt(problems_name[problems.index(problem)], pops)
    return hv_score, igd


def n_run(n, MODEL, problem, pf, reference_point, draw=False, s2f=False):
    '''
    运行n次
    :param n:
    :param problem:
    :return:
    '''
    igds = []
    hvs = []
    for i in range(n):
        print(f'第{i+1}次运行')
        # if i==0:
        #     s2f=True
        # else:
        #     s2f=False
        hv, igd = problem_test(MODEL, problem, pf, reference_point=reference_point, draw=draw, s2f=s2f)
        hvs.append(hv)
        igds.append(igd)
    # print("avgIGD={},minIGD={}\navgHV={},minHV={}".format(
    #     sum(igds)/n, min(igds), sum(hvs)/n, min(hvs)
    # ))
    return igds, hvs

def selectPF(pname):
    '''
    根据不同benchmark选择pareto前沿
    '''
    if 'WFG' in pname:
        pf = get_pflist('pf_files/wfg-pf/{}.3D.pf'.format(pname))
    elif 'DTLZ' in pname:
        pf = get_pflist('pf_files/n10000/{}.txt'.format(pname))
    else:
        return None
    return pf

def models_test(nrun, draw, s2f=False):
    '''
    对多个算法模型进行计算
    '''
    res = []
    for MODEL in models:
        resi = problems_test(nrun, MODEL, draw, s2f=s2f)
        res.append(resi)
    return res

def problems_test2(nrun, draw, s2f=False):
    '''
    一系列函数问题的测试
    :return:
    '''
    f1 = openpyxl.load_workbook('results/excels/result.xlsx')
    table = f1['Sheet1']
    f2 = open('results/res1.txt', 'w+', encoding='utf-8')
    i = 0
    for id in range(7):
        print('DTLZ{} starting……'.format(id+1))
        igdss, hvss = [], []
        pf = get_pflist(f'pf_files/n10000/DTLZ{id+1}.txt')
        for j in range(len(models)):
            igds, hvs = n_run(nrun, models[j], problems[id], pf, draw=draw, s2f=s2f)
            igdss.append(igds)
            hvss.append(hvs)
            table.cell(row=i + 3, column=j * 2 + 3).value = np.mean(igds)
            table.cell(row=i + 3, column=j * 2 + 4).value = np.std(hvs)
        f2.write(str(igdss)+'\n')
        f2.write(str(hvss)+'\n')
        draw_box(hvss, f'DTLZ{id+1}' + ' HV', f'results/photos/box/DTLZ{id+1}_HV.png')
        draw_box(igdss, f'DTLZ{id+1}' + ' IGD', f'results/photos/box/DTLZ{id+1}_IGD.png')
        i+=1
    for id in range(9):
        print('WFG{} starting……'.format(id + 1))
        pf = get_pflist(f"pf_files/wfg-pf/WFG{id+1}.3D.pf")
        igdss, hvss = [], []
        for j in range(len(models)):
            igds, hvs = n_run(nrun, models[j], problems[id+7], pf, draw=draw, s2f=s2f)
            igdss.append(igds)
            hvss.append(hvs)
            table.cell(row=i + 3, column=j * 2 + 3).value = np.mean(igds)
            table.cell(row=i + 3, column=j * 2 + 4).value = np.std(hvs)
        f2.write(str(igdss) + '\n')
        f2.write(str(hvss) + '\n')
        draw_box(hvss, f'WFG{id+1} HV', f'results/photos/box/WFG{id+1}_HV.png')
        draw_box(igdss, f'WFG{id+1} IGD', f'results/photos/box/WFG{id+1}_IGD.png')
        i+=1
    f1.save('results/excels/result.xlsx')
    f1.close()
    f2.close()
    print('程序结束')


def problems_test3(nrun, draw=False, s2f=False):
    '''
    一系列函数问题的测试
    :return:
    '''
    f2 = open('results/res1.txt', 'w+', encoding='utf-8')
    i = 0
    for id in range(7):
        s = f'DTLZ{id+1}'
        print('DTLZ{} starting……'.format(id+1))
        pf = get_pflist(f'pf_files/n10000/DTLZ{id+1}.txt')
        reference_point = get_referencepoint(pf)
        for j in range(len(models)):
            igds, hvs = n_run(nrun, models[j], problems[id], pf, reference_point, draw=draw, s2f=s2f)
            s += str(igds)+str(hvs)+'''
'''+'\n'
        s += '\n'
        print(s)
        f2.write(s)
        i+=1
    for id in range(9):
        s = f'WFG{id+1}'
        print('WFG{} starting……'.format(id + 1))
        pf = get_pflist(f"pf_files/wfg-pf/WFG{id+1}.3D.pf")
        reference_point = get_referencepoint(pf)
        for j in range(len(models)):
            igds, hvs = n_run(nrun, models[j], problems[id+7], pf, reference_point, draw=draw, s2f=s2f)
            s += str(igds) + str(hvs) + '''
''' + '\n'
        s += '\n'
        print(s)
        f2.write(s)
        i+=1
    f2.close()
    print('程序结束')

if __name__ == '__main__':
    # pf = get_pflist("pf_files/wfg-pf/WFG1.3D.pf")
    # pf = get_pflist('pf_files/n10000/{}.txt'.format('DTLZ6'))
    # reference_point = get_referencepoint2('WFG1')
    # problem_test(MOEADCODE, DTLZ6, pf, draw=False, s2f=False)
    # igdss, hvss = problems_test(False, s2f=True)
    # problem_test(DTLZ4, s2f=False)
    # igds, hvs = n_run(20, MOEAD, problems[7], pf, reference_point, False, s2f=False)
    # print('igds(std, mean):',np.std(igds), np.mean(igds))
    # print('hvs(std, mean):',np.std(hvs), np.mean(hvs))
    # print(igds)
    # print(hvs)
    # res = problems_test(1, MOEAD, False, s2f=False)
    # saveRes2Excel("./results/excels/{}.xls".format('moea-d-sade'), res)

    # res = models_test(40, False, s2f=False)
    # saveArray2Excel('./results/excels/result.xlsx', res)
    # write2file(res, 'results/res.txt')
    # draw_box2(res)
    # print('运行结束')

    problems_test3(1, True, True)

