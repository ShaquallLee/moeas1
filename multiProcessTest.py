# 使用多进程的方式运行程序
import multiprocessing
import numpy as np
import openpyxl
import time

from config import *
from test import n_run
from utils.drawResults import draw_box
from utils.pfget import get_pflist
from utils.referencePoint import get_referencepoint, get_referencepoint1, get_referencepoint2


def call_back(res):
    '''
    callback 返回调用函数，接收多进程返回的结果，进行处理
    '''
    # f1 = openpyxl.load_workbook('results/excels/result.xlsx')
    # table1 = f1['IGD']
    # table2 = f1['HV']
    # for j in range(len(models)):
    #     table1.cell(row=ri + 3, column=j * 2 + 3).value = np.mean(res[1][j])
    #     table1.cell(row=ri + 3, column=j * 2 + 4).value = np.std(res[1][j])
    #     table2.cell(row=ri + 3, column=j * 2 + 3).value = np.mean(res[2][j])
    #     table2.cell(row=ri + 3, column=j * 2 + 4).value = np.std(res[2][j])
    # f1.save('results/excels/result.xlsx')
    # f1.close()
    # print('excel写入成功')
    with open('results/res.txt', 'a', encoding='utf-8') as f:
        f.write(str(res))
    print('txt写入成功')


def prog_cell(ri, pf, bm_name, reference_point):
    '''
    单个进程单元
    '''
    print(bm_name+' starting……')
    res = []
    igdss, hvss = [], []
    for j in range(len(models)):
        print(f'model{j+1} starting……')
        igds, hvs = n_run(20, models[j], problems[ri], pf, reference_point)
        igdss.append(igds)
        hvss.append(hvs)
    res.append(bm_name)
    res.append(igdss)
    res.append(hvss)
    draw_box(hvss, f'{bm_name} HV', f'results/photos/box/{bm_name}_HV.png')
    draw_box(igdss, f'{bm_name} IGD', f'results/photos/box/{bm_name}_IGD.png')
    print(bm_name+' ending……')
    return res#,ri


def run_multiprocess():
    '''
    多进程主程序
    '''
    start = time.time()
    pool = multiprocessing.Pool(processes=16)
    for i in range(len(problems)):
        if i<7:
            problem_name = f'DTLZ{i+1}'
            pf = get_pflist(f'pf_files/n10000/{problem_name}.txt')
        else:
            problem_name = f'WFG{i-6}'
            pf = get_pflist(f"pf_files/wfg-pf/{problem_name}.3D.pf")
        reference_point = get_referencepoint(pf)
        pool.apply_async(prog_cell, (i, pf, problem_name, reference_point,), callback=call_back) #fei阻塞的
    pool.close()
    pool.join()  # 调用join之前，先调用close函数，否则会出错。执行完close后不会有新的进程加入到pool,join函数等待所有子进程结束
    print(f'运行结束,共用时{time.time()-start}')

if __name__ == '__main__':
    run_multiprocess()
    # pf = get_pflist('pf_files/n10000/{}.txt'.format('DTLZ1'))
    # reference_point = get_referencepoint(pf)
    # res, ri = prog_cell(5,pf,'DTLZ1', reference_point)
    # call_back(res,5)