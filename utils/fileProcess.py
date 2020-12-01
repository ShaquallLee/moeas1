#!/usr/bin/env python
# encoding: utf-8
# @author: lishaogang
# @file: fileProcess.py
# @time: 2020/10/22 0022 17:03
# @desc:

from xlwt import *
import time
import numpy as np
import xlrd
import openpyxl
import re
from config import problems_name

def readResFromTxt(file_name):
    with open(file_name, 'r', encoding='utf-8') as f:
        res = f.readline()
    match_text = '([A-Z]{3,4}\d)\[\[([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*)\], \[([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*)\], \[([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*)\], \[([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*)\]\]\[\[([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*)\], \[([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*)\], \[([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9e\-]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*)\], \[([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*), ([0-9]{1,}[.][0-9e\-]*)\]\]'
    match_res = re.findall(match_text, res)
    res_d = {}
    for i in range(len(match_res)):
        igd_l = []
        hv_l = []
        for j in range(4):
            igd_l.append(list(map(float,match_res[i][1+j*20: 1+(j+1)*20])))
            hv_l.append(list(map(float,match_res[i][81+j*20: 81+(j+1)*20])))
        res_d[match_res[i][0]] = [igd_l, hv_l]
    return res_d

def saveArray2Excel(name, data):
    '''
    保存数据到excel文件中
    :param name:
    :param data:
    :return:
    '''
    f = openpyxl.load_workbook(name)
    table = f['Sheet1']
    i = 0
    for proname in problems_name:
        value = data.get(proname)
        if value!=None:
            igd, hv = value
            for j in range(4):
                table.cell(row=i+3, column=j*2+3).value = np.mean(igd)
                table.cell(row=i+3, column=j*2+4).value = np.std(hv)
        i+=1
    f.save(name)
    f.close()
    print("保存数据到{}文件成功".format(name))

def saveRes2Excel(fname, res):
    '''
    保存数据到excel文件中
    :param name:
    :param data:
    :return:
    '''
    f = Workbook(encoding='utf-8')
    for name, data in res.items():
        table = f.add_sheet(name)
        for i in range(len(data[0])):
            table.write(0, i+1, "第{}次".format(i+1))
        table.write(0, len(data[0])+1, '均值')
        table.write(0, len(data[0])+2, '最小值')
        table.write(0, len(data[0])+3, '最大值')
        table.write(1,0, "IGD")
        table.write(2,0, "HV")
        for i in range(len(data)):
            for j in range(len(data[i])):
                table.write(i+1,j+1, data[i][j])
            table.write(i+1,len(data[i])+1, sum(data[i])/len(data[i]))
            table.write(i+1,len(data[i])+2, min(data[i]))
            table.write(i+1,len(data[i])+3, max(data[i]))
    f.save(fname)
    print("保存数据到{}文件成功".format(fname))

def savePareto2Txt(name,pareto):
    '''
    保存Pareto前沿到txt文件中
    :param name:
    :param pareto:
    :return:
    '''
    tname = './results/{}_res_{}.txt'.format(name, time.time())
    with open(tname, 'w+', encoding='utf-8') as f:
        for items in pareto:
            if items!=[]:
                f.write('\t'.join(list(map(str, list(items))))+'\n')
            else:
                print('EORRO:88888888')
                print(pareto)
    print("保存至‘{}'成功".format(tname))

def readPareto4Txt(name):
    '''
    从txt文件中读取pareto前沿
    :param name:
    :return:
    '''
    pareto = []
    with open('./results/{}.txt'.format(name),'r', encoding='utf-8') as f:
        lines = f.readlines()
        for line in lines:
            pareto.append(list(map(float, line.strip().split('\t'))))
    print("读取‘./results/{}_res.txt'成功".format(name))
    return pareto

def write2file(data, name):
    '''
    将数据存储到txt文件中
    '''
    try:
        with open(f'results/{name}.txt', 'w+', encoding='utf-8') as f:
            f.write(str(data))
            print('成功保存文件')
    except:
        print('保存结果到txt文件失败')

if __name__ == '__main__':
    res = readResFromTxt('../results/res.txt')
    saveArray2Excel('../results/excels/result.xlsx', res)
